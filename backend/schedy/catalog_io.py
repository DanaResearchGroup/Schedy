"""Catalog <-> CSV/Excel import & export.

The catalog is the department's hand-built source of truth, so it must be
savable to a file (to version, back up, share) and reloadable. The format is one
row per course; list-valued fields (programs, lab_days, lecturer_ids, ta_ids) use
';' *inside* the cell so the file stays plainly comma-delimited and opens cleanly
in Excel. Pure and testable: formatting/parsing operate on text and records, not
files.
"""

from __future__ import annotations

import csv
import io

from .catalog import Course
from .domain import CourseRole, Program

# Column order of the CSV/Excel file. This header IS the format spec.
COLUMNS = [
    "number", "name_he", "name_en", "programs", "year", "role",
    "lecture_boxes", "num_exercise_groups", "exercise_boxes", "lab_boxes", "lab_days",
    "expected_enrollment", "needs_computer_farm", "is_remote",
    "is_external", "ext_day", "ext_start_min", "ext_end_min", "ext_room",
    "lecturer_ids", "ta_ids",
]


def _join(xs) -> str:
    return ";".join(str(x) for x in xs)


def _bool(b) -> str:
    return "true" if b else "false"


def _course_to_row(c: Course) -> dict:
    return {
        "number": c.number, "name_he": c.name_he, "name_en": c.name_en,
        "programs": _join(p.value for p in c.programs), "year": c.year,
        "role": c.role.value,
        "lecture_boxes": c.lecture_boxes, "num_exercise_groups": c.num_exercise_groups,
        "exercise_boxes": c.exercise_boxes, "lab_boxes": c.lab_boxes,
        "lab_days": _join(c.lab_days),
        "expected_enrollment": c.expected_enrollment,
        "needs_computer_farm": _bool(c.needs_computer_farm),
        "is_remote": _bool(c.is_remote),
        "is_external": _bool(c.is_external),
        "ext_day": "" if c.ext_day is None else c.ext_day,
        "ext_start_min": "" if c.ext_start_min is None else c.ext_start_min,
        "ext_end_min": "" if c.ext_end_min is None else c.ext_end_min,
        "ext_room": c.ext_room or "",
        "lecturer_ids": _join(c.lecturer_ids), "ta_ids": _join(c.ta_ids),
    }


def to_csv(courses) -> str:
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=COLUMNS)
    w.writeheader()
    for c in courses:
        w.writerow(_course_to_row(c))
    return buf.getvalue()


# --- parsing ------------------------------------------------------------- #

def _split(s) -> list[str]:
    return [x.strip() for x in str(s or "").split(";") if x.strip()]


def _int(s, default: int = 0) -> int:
    s = str(s or "").strip()
    return int(float(s)) if s else default


def _opt_int(s):
    s = str(s or "").strip()
    return int(float(s)) if s else None


def _truthy(s) -> bool:
    return str(s).strip().lower() in ("1", "true", "yes", "y", "t", "כן")


def row_to_course(row: dict) -> Course | None:
    """Build a Course from one record; returns None for a number-less row."""
    number = str(row.get("number", "")).strip()
    if not number:
        return None
    role_raw = str(row.get("role") or "core").strip().lower()
    return Course(
        number=number,
        name_he=str(row.get("name_he") or "").strip(),
        name_en=str(row.get("name_en") or "").strip(),
        programs=[Program(p) for p in _split(row.get("programs"))],
        year=_int(row.get("year"), 1),
        role=CourseRole(role_raw) if role_raw else CourseRole.CORE,
        lecture_boxes=_int(row.get("lecture_boxes")),
        num_exercise_groups=_int(row.get("num_exercise_groups")),
        exercise_boxes=_int(row.get("exercise_boxes"), 1),
        lab_boxes=_int(row.get("lab_boxes")),
        lab_days=[int(x) for x in _split(row.get("lab_days"))],
        expected_enrollment=_int(row.get("expected_enrollment")),
        needs_computer_farm=_truthy(row.get("needs_computer_farm")),
        is_remote=_truthy(row.get("is_remote")),
        is_external=_truthy(row.get("is_external")),
        ext_day=_opt_int(row.get("ext_day")),
        ext_start_min=_opt_int(row.get("ext_start_min")),
        ext_end_min=_opt_int(row.get("ext_end_min")),
        ext_room=str(row.get("ext_room") or "").strip() or None,
        lecturer_ids=_split(row.get("lecturer_ids")),
        ta_ids=_split(row.get("ta_ids")),
    )


def from_rows(rows) -> list[Course]:
    out = []
    for row in rows:
        c = row_to_course(row)
        if c:
            out.append(c)
    return out


def from_csv(text: str) -> list[Course]:
    text = text.lstrip("﻿")  # tolerate the UTF-8 BOM Excel writes
    return from_rows(csv.DictReader(io.StringIO(text)))


def from_xlsx_bytes(data: bytes) -> list[Course]:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    if not rows:
        return []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    records = [
        {header[i]: ("" if v is None else v) for i, v in enumerate(r) if i < len(header)}
        for r in rows[1:]
    ]
    return from_rows(records)


def template_csv() -> str:
    """A documented example file: the header plus two illustrative courses."""
    examples = [
        Course(
            number="00540315", name_he="תרמודינמיקה א׳", name_en="Thermodynamics A",
            programs=[Program.CHEME], year=2, role=CourseRole.CORE,
            lecture_boxes=2, num_exercise_groups=2, exercise_boxes=1,
            lab_boxes=2, lab_days=[0, 3], expected_enrollment=70,
            lecturer_ids=["prof_bar"], ta_ids=["ta_hedva", "ta_yossi"],
        ),
        Course(
            number="01040031", name_he="חשבון 1", name_en="Calculus 1 (external)",
            programs=[Program.CHEME, Program.BIOCHEME], year=1, role=CourseRole.CORE,
            is_external=True, ext_day=1, ext_start_min=510, ext_end_min=630,
            ext_room="Math building",
        ),
    ]
    return to_csv(examples)
