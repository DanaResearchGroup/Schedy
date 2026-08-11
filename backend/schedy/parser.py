"""Skeleton Parser (mostly pure).

Turns the Technion "skeleton" XLSX into clean OfferedSession objects, filtered to
the department's relevant course numbers. Columns are located by their Hebrew
header text (not fixed indices) so the parser survives column reordering.

The file-touching part is a thin shell (`parse_skeleton`); all the real logic is
in `parse_rows`, a pure function over header + row lists, so it is tested without
any spreadsheet.

Columns the scheduler reasons about get their own `OfferedSession` field. Every
other named column of the export is carried verbatim in `details`, so the review
screen can show a course's whole record without this module growing a field per
column. See `examples/README.md` for the format as a whole.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime

from .domain import SessionType

# Hebrew header labels -> internal field keys.
_HEADERS = {
    "course_number": "מקצוע",
    "name_he": "תיאור מקצוע עברית",
    "name_en": "תיאור מקצוע אנגלית",
    "package": "תיאור חבילת רישום",
    "event_type": "סוג אירוע D",
    "room": "חדר",
    "faculty": "פקולטה",
    "language": "שפת הוראת אירוע",
    "person": "אדם מוקצה",
}

# Everything else the export names, kept verbatim for the review screen. Order
# here is the order the details are displayed in.
#
# Two columns are deliberately absent: "אדם מוקצה (מספר עובד)" and
# "אדם מוקצה (ת.ז.)" — an employee number and a national ID. The person's *name*
# is already captured above; the identifiers carry no scheduling meaning, so they
# are dropped here rather than persisted into the app's database.
_DETAIL_HEADERS = {
    "building": "בניין",
    "academic_level": "רמה אקדמית",
    "course_language": "שפת הוראת מקצוע",
    "weekly_hours": "שעות הוראה בשבוע",
    "central_planning": "תכנון מרכזי",
    "room_approval_status": "סטאטוס אישור חדר",
    "registered_ug": "מספר רשומים UG",
    "requests_ug": "מספר בקשות רישום UG",
    "waitlist_ug": "מספר רשימת המתנה UG",
    "registered_ug_total": 'מספר רשומים UG – סה"כ',
    "requests_ug_total": 'מספר בקשות רישום UG – סה"כ',
    "waitlist_ug_total": 'מספר רשימת המתנה UG – סה"כ',
    "registered_gr": "מספר רשומים GR",
    "requests_gr": "מספר בקשות רישום GR",
    "registered_gr_total": 'מספר רשומים GR – סה"כ',
    "requests_gr_total": 'מספר בקשות רישום GR – סה"כ',
    "grad_package_capacity": "קיבולת חבילת רישום מוסמכים",
    "exam_a_date": "תאריך מועד א",
    "exam_b_date": "תאריך מועד ב",
    "quiz_date": "תאריך בחן",
    "show_in_catalog": "הצגת מקצוע בקטלוג",
    "course_in_tens": "מקצוע בעשרה",
    "semester_note": "הערה לסמסטר",
    "semester_note_2": "הערה לסמסטר2",
    "semester_note_3": "הערה לסמסטר3",
}

_HEB_DAYS = {
    "ראשון": 0, "שני": 1, "שלישי": 2, "רביעי": 3,
    "חמישי": 4, "שישי": 5, "שבת": 6,
}

_EVENT_TYPES = {
    "הרצאה": SessionType.LECTURE,
    "תרגול": SessionType.EXERCISE,
    "תרגיל": SessionType.EXERCISE,
    "מעבדה": SessionType.LAB,
}

_TIME_RE = re.compile(r"(\d{1,2}):(\d{2})\s*-\s*(\d{1,2}):(\d{2})")


@dataclass(frozen=True)
class OfferedSession:
    course_number: str
    name_he: str
    name_en: str
    event_type: SessionType | None
    group_code: str | None        # leading code-like token, e.g. "SE011"
    package: str                  # full registration-package text (for matching)
    day: int | None               # 0=Sun .. 6=Sat, or None if no time slot
    start_min: int | None
    end_min: int | None
    room: str
    faculty: str
    language: str
    person: str
    row: int                      # 1-based source row, for the review screen
    # Every other named column, slug -> text. Blank cells are left out, so an
    # absent key and an empty one mean the same thing: the export said nothing.
    details: dict[str, str] = field(default_factory=dict)


def _norm(s) -> str:
    if s is None:
        return ""
    if isinstance(s, datetime):
        s = s.date()
    if isinstance(s, date):
        return s.isoformat()
    return str(s).strip()


# Header text as written varies in ways that carry no meaning: the export uses an
# en-dash in some labels, a Hebrew gershayim in others, and stray double spaces.
# Comparing canonical forms keeps a hand-typed header matching a real one.
_DASHES = str.maketrans({"–": "-", "—": "-", "־": "-"})
_QUOTES = str.maketrans({"״": '"', "׳": "'", "“": '"', "”": '"', "’": "'"})


def _canon(label: str) -> str:
    return " ".join(label.translate(_DASHES).translate(_QUOTES).split())


_HEADER_BY_LABEL = {_canon(heb): key for key, heb in _HEADERS.items()}
_DETAIL_BY_LABEL = {_canon(heb): key for key, heb in _DETAIL_HEADERS.items()}


def _extract_group_code(package: str) -> str | None:
    """Leading token of the package text if it looks like a group code."""
    token = package.strip().split(" ", 1)[0] if package.strip() else ""
    return token if any(ch.isdigit() for ch in token) else None


def _parse_time(cell: str) -> tuple[int, int] | None:
    m = _TIME_RE.search(cell)
    if not m:
        return None
    h1, m1, h2, m2 = (int(x) for x in m.groups())
    return h1 * 60 + m1, h2 * 60 + m2


def parse_rows(
    header: list, rows: list[list], relevant_course_numbers: set[str] | None = None
) -> list[OfferedSession]:
    """Pure core: header row + data rows -> filtered OfferedSession list."""
    idx = {key: None for key in _HEADERS}
    detail_cols: dict[str, int] = {}
    day_cols: dict[int, int] = {}
    for col, raw in enumerate(header):
        label = _canon(_norm(raw))
        if label in _HEADER_BY_LABEL:
            idx[_HEADER_BY_LABEL[label]] = col
        elif label in _DETAIL_BY_LABEL:
            detail_cols[_DETAIL_BY_LABEL[label]] = col
        if label in _HEB_DAYS:
            day_cols[_HEB_DAYS[label]] = col

    if idx["course_number"] is None:
        # A visual draft timetable (day headers across the top, time ranges down
        # the first column, free-text course names in the cells) has no column
        # structure at all — detect it and say so, rather than a cryptic
        # "missing column". Signals: >=2 day-name header cells, and/or time
        # ranges in the first column of the data rows.
        day_headers = sum(1 for raw in header if _norm(raw) in _HEB_DAYS)
        timed_first_col = sum(
            1 for row in rows[:12]
            if row and _parse_time(_norm(row[0]))
        )
        if day_headers >= 2 or timed_first_col >= 2:
            raise ValueError(
                "this looks like a visual weekly timetable grid (day headers + "
                "time rows), not a Technion registration export; upload the "
                "registration XLSX (with a 'מקצוע' course-number column)"
            )
        raise ValueError("skeleton header missing course-number column ('מקצוע')")

    def get(row, key) -> str:
        c = idx[key]
        return _norm(row[c]) if c is not None and c < len(row) else ""

    out: list[OfferedSession] = []
    for r, row in enumerate(rows, start=2):  # +1 header, +1 to 1-base
        course = get(row, "course_number")
        if not course:
            continue
        if relevant_course_numbers is not None and course not in relevant_course_numbers:
            continue

        day = start = end = None
        for d, col in day_cols.items():
            cell = _norm(row[col]) if col < len(row) else ""
            t = _parse_time(cell)
            if t:
                day, (start, end) = d, t
                break

        # Detail order follows _DETAIL_HEADERS, not the file's column order, so
        # the review screen reads the same way whatever the export shuffles.
        details = {}
        for key in _DETAIL_HEADERS:
            col = detail_cols.get(key)
            if col is None or col >= len(row):
                continue
            value = _norm(row[col])
            if value:
                details[key] = value

        package = get(row, "package")
        out.append(OfferedSession(
            course_number=course,
            name_he=get(row, "name_he"),
            name_en=get(row, "name_en"),
            event_type=_EVENT_TYPES.get(get(row, "event_type")),
            group_code=_extract_group_code(package),
            package=package,
            day=day,
            start_min=start,
            end_min=end,
            room=get(row, "room"),
            faculty=get(row, "faculty"),
            language=get(row, "language"),
            person=get(row, "person"),
            row=r,
            details=details,
        ))
    return out


def parse_skeleton(
    path: str, relevant_course_numbers: set[str] | None = None, sheet: str | None = None
) -> list[OfferedSession]:
    """Load a Technion skeleton XLSX and return filtered OfferedSession objects."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    all_rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    if not all_rows:
        return []
    return parse_rows(all_rows[0], all_rows[1:], relevant_course_numbers)
