"""Courses-of-interest file <-> records.

The department cares about a few dozen of the university's thousands of courses,
and that set barely moves from year to year. So it lives in a small file the
planner maintains by hand, and that file is what filters the imported skeleton.

Being hand-maintained, the format has to be forgiving. Two shapes are accepted:

    number,name              <- headed, either column order, Hebrew or English
    00540315,תרמודינמיקה א׳

    00540315                 <- bare list, no header at all
    00540319

Pure and testable: everything below operates on text and records, with the two
file shells (`from_csv`, `from_xlsx_bytes`) delegating straight into them.
"""

from __future__ import annotations

import csv
import io

from . import uploads

# A course record is a plain dict {"number": str, "name": str} — the same shape
# the /courses-of-interest endpoint has always stored.
COLUMNS = ["number", "name"]

# Header cells that name each column. Compared case-folded and stripped.
_NUMBER_ALIASES = {
    "number", "course_number", "course", "course number", "no", "#",
    "מקצוע", "מספר", "מספר מקצוע",
}
_NAME_ALIASES = {
    "name", "name_he", "name_en", "course_name", "title", "description",
    "שם", "שם מקצוע", "תיאור", "תיאור מקצוע", "תיאור מקצוע עברית",
}

# Technion course numbers are eight digits, written with their leading zeros.
NUMBER_WIDTH = 8


def _cell(v) -> str:
    """One cell as trimmed text, undoing the two ways a spreadsheet mangles it."""
    if v is None:
        return ""
    # openpyxl hands back numeric-looking cells as int/float, so "00540315"
    # arrives as 540315 and "540315.0" is a float's repr. Both must survive.
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v).strip()


def normalize_number(v) -> str:
    """Course number as the skeleton writes it: eight digits, zeros kept.

    Excel drops the leading zeros of `00540315` the moment the file is opened and
    saved. Restoring them here is what lets an Excel-round-tripped file still
    match the skeleton; without it the interest list silently matches nothing.
    """
    s = _cell(v)
    if s.isdigit() and len(s) < NUMBER_WIDTH:
        return s.zfill(NUMBER_WIDTH)
    return s


def _is_number_like(s: str) -> bool:
    return bool(s) and s.isdigit()


# --- writing ------------------------------------------------------------- #

def to_csv(items) -> str:
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=COLUMNS)
    w.writeheader()
    for it in items:
        w.writerow({"number": it.get("number", ""), "name": it.get("name", "")})
    return buf.getvalue()


def template_csv() -> str:
    """A documented example file: the header plus a few illustrative courses."""
    return to_csv([
        {"number": "00540315", "name": "תרמודינמיקה א׳"},
        {"number": "00540319", "name": "תופעות מעבר"},
        {"number": "01250300", "name": "כימיה אורגנית"},
    ])


# --- parsing ------------------------------------------------------------- #

def _column_map(header: list[str]) -> tuple[int, int | None] | None:
    """Locate (number_col, name_col) in a header row, or None if it isn't one."""
    number_col = name_col = None
    for i, cell in enumerate(header):
        key = cell.strip().casefold()
        if number_col is None and key in _NUMBER_ALIASES:
            number_col = i
        elif name_col is None and key in _NAME_ALIASES:
            name_col = i
    return None if number_col is None else (number_col, name_col)


def from_rows(rows: list[list]) -> list[dict]:
    """Core: a list of raw cell rows -> deduplicated course records.

    Detects the headed and bare shapes, then normalises numbers and drops blanks.
    """
    grid = [[_cell(c) for c in row] for row in rows]
    grid = [row for row in grid if any(row)]  # blank lines are just spacing
    if not grid:
        return []

    cols = _column_map(grid[0])
    if cols is not None:
        number_col, name_col = cols
        body = grid[1:]
    elif _is_number_like(grid[0][0]):
        # No recognised header and the first cell reads like a course number:
        # the whole file is data. Treat a second column, if present, as the name.
        number_col, name_col = 0, (1 if len(grid[0]) > 1 else None)
        body = grid
    else:
        # A header we don't recognise. Fall back to positional columns rather
        # than returning nothing, but skip that first row.
        number_col, name_col = 0, (1 if len(grid[0]) > 1 else None)
        body = grid[1:]

    out: list[dict] = []
    seen: set[str] = set()
    for row in body:
        if number_col >= len(row):
            continue
        number = normalize_number(row[number_col])
        if not number or number in seen:
            continue
        seen.add(number)
        name = row[name_col] if name_col is not None and name_col < len(row) else ""
        out.append({"number": number, "name": name})
    return out


def from_csv(text: str) -> list[dict]:
    text = text.lstrip("﻿")  # tolerate the UTF-8 BOM Excel writes
    return from_rows(list(csv.reader(io.StringIO(text))))


def from_xlsx_bytes(data: bytes) -> list[dict]:
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return from_rows(rows)


def from_upload(data: bytes, filename: str) -> list[dict]:
    """Parse an uploaded interest file, choosing the reader by extension."""
    uploads.reject_legacy_xls(filename)
    if uploads.is_excel(filename):
        return from_xlsx_bytes(data)
    return from_csv(data.decode("utf-8-sig"))
